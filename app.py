from __future__ import annotations

import calendar
import html
import io
from dataclasses import dataclass
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_TITLE = "老健向けAIシフト作成システム"
SHIFT_ORDER = ["早番", "日勤", "遅番", "夜勤"]
SHIFT_TIMES = {
    "早番": "7:00〜16:00",
    "日勤": "8:30〜17:30",
    "遅番": "10:30〜19:30",
    "夜勤": "16:30〜翌9:30",
    "明け": "—",
    "休み": "—",
}
REQUIRED = {
    "看護師": {"早番": 1, "日勤": 4, "遅番": 1, "夜勤": 1},
    "介護士": {"早番": 4, "日勤": 8, "遅番": 4, "夜勤": 4},
}
SHIFT_COLORS = {
    "早番": "#DBEAFE",
    "日勤": "#DCFCE7",
    "遅番": "#FEF3C7",
    "夜勤": "#EDE9FE",
    "明け": "#FCE7F3",
    "休み": "#F1F5F9",
}
WEEKDAYS = "月火水木金土日"


@dataclass
class GenerationResult:
    staff: pd.DataFrame
    schedule: pd.DataFrame
    daily: pd.DataFrame
    summary: pd.DataFrame
    warnings: pd.DataFrame


def make_staff() -> pd.DataFrame:
    """営業デモ用の、毎回同じダミー職員を生成する。"""
    nurse_names = [
        "佐藤 美咲", "鈴木 陽子", "高橋 直子", "田中 智子", "伊藤 由美",
        "渡辺 麻衣", "山本 香織", "中村 明日香", "小林 典子", "加藤 奈緒",
    ]
    care_names = [
        "吉田 健太", "山田 翔", "佐々木 彩", "山口 大輔", "松本 和也", "井上 亮",
        "木村 優", "林 拓也", "清水 愛", "斎藤 誠", "山崎 遥", "森 俊介",
        "池田 葵", "橋本 達也", "阿部 菜月", "石川 悠斗", "前田 朱里", "藤田 蓮",
        "後藤 陽介", "岡田 結衣", "長谷川 陸", "村上 七海", "近藤 湊", "石井 杏奈",
        "坂本 蒼", "遠藤 美月", "青木 樹", "藤井 花",
    ]
    rows = []
    for i, name in enumerate(nurse_names, 1):
        rows.append({"職員ID": f"N{i:02d}", "職種": "看護師", "氏名": name})
    for i, name in enumerate(care_names, 1):
        rows.append({"職員ID": f"C{i:02d}", "職種": "介護士", "氏名": name})
    return pd.DataFrame(rows)


def _work_streak(assignments: list[str]) -> int:
    streak = 0
    for value in reversed(assignments):
        if value in SHIFT_ORDER:
            streak += 1
        else:
            break
    return streak


def _generate_for_role(
    role_staff: pd.DataFrame, days: list[date], required: dict[str, int], seed: int
) -> dict[str, list[str]]:
    """必要人数を必ず満たす循環型スケジューラ。

    夜勤人数単位のグループを作り、夜勤→明け→休みを周期化する。
    残りの勤務可能者数が日中帯の必要人数と一致するため、欠員を
    発生させず、最大連勤も5日以内になる。
    """
    ids = role_staff["職員ID"].tolist()
    assignments = {staff_id: [] for staff_id in ids}
    counts = {staff_id: {shift: 0 for shift in SHIFT_ORDER} for staff_id in ids}
    weekend_work = {staff_id: 0 for staff_id in ids}

    night_size = required["夜勤"]
    total_required = sum(required.values())
    if len(ids) % night_size != 0:
        raise ValueError("職員数は夜勤必要人数の倍数にしてください。")
    cycle = len(ids) // night_size
    working_groups = total_required // night_size
    off_groups = cycle - working_groups
    if total_required % night_size or off_groups < 2:
        raise ValueError("現在の人数構成では循環シフトを作成できません。")

    groups = [ids[index:index + night_size] for index in range(0, len(ids), night_size)]
    group_of = {staff_id: group_index for group_index, group in enumerate(groups) for staff_id in group}
    # phase 0=夜勤、1=明け、2=休み。余剰人員分の休日は周期後半へ置く。
    extra_off_phases = {cycle - 2 * index for index in range(1, off_groups - 1)}

    for day_index, current_day in enumerate(days):
        today: dict[str, str | None] = {}
        for staff_id in ids:
            phase = (day_index - group_of[staff_id]) % cycle
            if phase == 0:
                today[staff_id] = "夜勤"
                counts[staff_id]["夜勤"] += 1
            elif phase == 1:
                today[staff_id] = "明け"
            elif phase == 2 or phase in extra_off_phases:
                today[staff_id] = "休み"
            else:
                today[staff_id] = None

        available = [staff_id for staff_id in ids if today[staff_id] is None]
        for shift in ["早番", "遅番", "日勤"]:
            for _ in range(required[shift]):
                selected = min(
                    available,
                    key=lambda staff_id: (
                        counts[staff_id][shift] * 10,
                        weekend_work[staff_id] if current_day.weekday() >= 5 else 0,
                        sum(counts[staff_id].values()),
                        staff_id,
                    ),
                )
                today[selected] = shift
                available.remove(selected)
                counts[selected][shift] += 1

        # 理論上availableは空。将来職員数を増やした場合は余剰を休みにする。
        for staff_id in ids:
            value = today[staff_id] or "休み"
            assignments[staff_id].append(value)
            if current_day.weekday() >= 5 and value in SHIFT_ORDER:
                weekend_work[staff_id] += 1
    return assignments


def _build_daily(schedule: pd.DataFrame, staff: pd.DataFrame, days: list[date]) -> pd.DataFrame:
    role_map = staff.set_index("職員ID")["職種"].to_dict()
    rows = []
    for day in days:
        column = day.strftime("%Y-%m-%d")
        for role in ["看護師", "介護士"]:
            role_ids = [staff_id for staff_id, value in role_map.items() if value == role]
            for shift in SHIFT_ORDER:
                actual = int((schedule.loc[role_ids, column] == shift).sum())
                needed = REQUIRED[role][shift]
                rows.append({
                    "日付": f"{day.month}/{day.day}({WEEKDAYS[day.weekday()]})",
                    "職種": role,
                    "勤務区分": shift,
                    "必要人数": needed,
                    "配置人数": actual,
                    "差分": actual - needed,
                    "判定": "充足" if actual >= needed else "不足",
                })
    return pd.DataFrame(rows)


def _build_summary(schedule: pd.DataFrame, staff: pd.DataFrame, days: list[date]) -> pd.DataFrame:
    weekend_columns = [day.strftime("%Y-%m-%d") for day in days if day.weekday() >= 5]
    rows = []
    for _, person in staff.iterrows():
        values = schedule.loc[person["職員ID"]]
        row = {"職員ID": person["職員ID"], "職種": person["職種"], "氏名": person["氏名"]}
        for shift in SHIFT_ORDER + ["明け", "休み"]:
            row[f"{shift}回数"] = int((values == shift).sum())
        row["勤務日数"] = sum(row[f"{shift}回数"] for shift in SHIFT_ORDER)
        row["土日勤務日数"] = int(values[weekend_columns].isin(SHIFT_ORDER).sum()) if weekend_columns else 0
        rows.append(row)
    return pd.DataFrame(rows)


def _build_daily_roster(schedule: pd.DataFrame, staff: pd.DataFrame, days: list[date]) -> pd.DataFrame:
    """日付・勤務区分ごとの職種別氏名一覧を作る。"""
    staff_by_id = staff.set_index("職員ID")
    rows = []
    for day in days:
        column = day.strftime("%Y-%m-%d")
        for shift in SHIFT_ORDER + ["明け", "休み"]:
            nurse_ids = [
                staff_id for staff_id in schedule.index
                if schedule.at[staff_id, column] == shift and staff_by_id.at[staff_id, "職種"] == "看護師"
            ]
            care_ids = [
                staff_id for staff_id in schedule.index
                if schedule.at[staff_id, column] == shift and staff_by_id.at[staff_id, "職種"] == "介護士"
            ]
            nurse_names = [staff_by_id.at[staff_id, "氏名"] for staff_id in nurse_ids]
            care_names = [staff_by_id.at[staff_id, "氏名"] for staff_id in care_ids]
            if shift in SHIFT_ORDER:
                nurse_needed = REQUIRED["看護師"][shift]
                care_needed = REQUIRED["介護士"][shift]
                fulfilled = len(nurse_names) >= nurse_needed and len(care_names) >= care_needed
                needed_text = f"看護師{nurse_needed}名・介護士{care_needed}名"
                judgment = "充足" if fulfilled else "不足"
            else:
                needed_text = "基準対象外"
                judgment = "確認用"
            rows.append({
                "日付": f"{day.year}年{day.month}月{day.day}日（{WEEKDAYS[day.weekday()]}）",
                "勤務区分": shift,
                "看護師名": "、".join(nurse_names) if nurse_names else "—",
                "介護士名": "、".join(care_names) if care_names else "—",
                "必要人数": needed_text,
                "配置人数": f"看護師{len(nurse_names)}名・介護士{len(care_names)}名",
                "判定": judgment,
            })
    return pd.DataFrame(rows)


def _build_warnings(
    schedule: pd.DataFrame, staff: pd.DataFrame, daily: pd.DataFrame, days: list[date]
) -> pd.DataFrame:
    warnings = []
    shortages = daily[daily["差分"] < 0]
    for _, row in shortages.iterrows():
        warnings.append({
            "重要度": "不足",
            "対象": f"{row['日付']} {row['職種']}・{row['勤務区分']}",
            "内容": f"必要{row['必要人数']}名に対し{row['配置人数']}名（{abs(row['差分'])}名不足）",
        })

    name_map = staff.set_index("職員ID")["氏名"].to_dict()
    for staff_id, values in schedule.iterrows():
        series = values.tolist()
        if series.count("休み") < 9:
            warnings.append({
                "重要度": "注意", "対象": name_map[staff_id],
                "内容": f"充足を優先したため月休日が{series.count('休み')}日です（目標：9日以上）",
            })
        streak = 0
        for index, value in enumerate(series):
            streak = streak + 1 if value in SHIFT_ORDER else 0
            if streak > 5:
                warnings.append({
                    "重要度": "警告", "対象": name_map[staff_id],
                    "内容": f"{days[index].month}/{days[index].day}時点で連続勤務が{streak}日です",
                })
                break
        for index in range(len(series) - 1):
            if series[index] == "夜勤" and series[index + 1] != "明け":
                warnings.append({"重要度": "警告", "対象": name_map[staff_id], "内容": "夜勤翌日の明けが未設定です"})
        for index in range(len(series) - 1):
            if series[index] == "明け" and series[index + 1] != "休み":
                warnings.append({"重要度": "注意", "対象": name_map[staff_id], "内容": "明け翌日が休みではありません"})
    if not warnings:
        warnings.append({"重要度": "正常", "対象": "全体", "内容": "不足・警告はありません"})
    return pd.DataFrame(warnings)


@st.cache_data(show_spinner=False)
def generate_schedule(year: int, month: int) -> GenerationResult:
    staff = make_staff()
    day_count = calendar.monthrange(year, month)[1]
    days = [date(year, month, day) for day in range(1, day_count + 1)]
    combined: dict[str, list[str]] = {}
    for role, seed_offset in [("看護師", 11), ("介護士", 29)]:
        role_staff = staff[staff["職種"] == role]
        combined.update(_generate_for_role(role_staff, days, REQUIRED[role], year * 100 + month + seed_offset))

    columns = [day.strftime("%Y-%m-%d") for day in days]
    schedule = pd.DataFrame.from_dict(combined, orient="index", columns=columns)
    schedule.index.name = "職員ID"
    daily = _build_daily(schedule, staff, days)
    summary = _build_summary(schedule, staff, days)
    warnings = _build_warnings(schedule, staff, daily, days)
    return GenerationResult(staff, schedule, daily, summary, warnings)


def display_schedule(result: GenerationResult, year: int, month: int) -> pd.DataFrame:
    output = result.staff.set_index("職員ID")[["職種", "氏名"]].join(result.schedule)
    renamed = {}
    for column in result.schedule.columns:
        day = pd.Timestamp(column)
        renamed[column] = f"{day.day}\n({WEEKDAYS[day.weekday()]})"
    return output.rename(columns=renamed).reset_index(drop=True)


def style_schedule(frame: pd.DataFrame):
    def color_shift(value: object) -> str:
        color = SHIFT_COLORS.get(str(value))
        return f"background-color: {color}; color: #0F172A; font-weight: 600" if color else ""
    return frame.style.map(color_shift)


def make_excel(result: GenerationResult, year: int, month: int) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    navy = "1E3A5F"
    blue = "2563EB"
    light_blue = "DBEAFE"
    red = "DC2626"
    thin = Side(style="thin", color="CBD5E1")

    def prepare_sheet(title: str, headers: list[str]):
        ws = wb.create_sheet(title)
        ws.freeze_panes = "A3"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        cell = ws.cell(1, 1, f"{year}年{month}月 {title}")
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="left")
        for col, header in enumerate(headers, 1):
            h = ws.cell(2, col, header)
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor=blue)
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
        return ws

    schedule_view = display_schedule(result, year, month)
    ws = prepare_sheet("月間シフト表", list(schedule_view.columns))
    for row_index, row in enumerate(schedule_view.itertuples(index=False), 3):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin, right=thin)
            if str(value) in SHIFT_COLORS:
                cell.fill = PatternFill("solid", fgColor=SHIFT_COLORS[str(value)].replace("#", ""))
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 14
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.freeze_panes = "C3"

    day_count = calendar.monthrange(year, month)[1]
    days = [date(year, month, day) for day in range(1, day_count + 1)]
    daily_roster = _build_daily_roster(result.schedule, result.staff, days)
    sheets = [
        ("日別配置チェック表", result.daily),
        ("職員別集計表", result.summary),
        ("不足・警告一覧", result.warnings),
        ("日別出勤者一覧", daily_roster),
    ]
    for title, frame in sheets:
        sheet = prepare_sheet(title, list(frame.columns))
        for row_index, row in enumerate(frame.itertuples(index=False), 3):
            for col_index, value in enumerate(row, 1):
                cell = sheet.cell(row_index, col_index, value)
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if value in ("不足", "警告") or (title == "日別配置チェック表" and frame.columns[col_index - 1] == "差分" and isinstance(value, int) and value < 0):
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                    cell.font = Font(color=red, bold=True)
                elif row_index % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="F8FAFC")
        for col_index, column in enumerate(frame.columns, 1):
            max_len = max([len(str(column))] + [len(str(value)) for value in frame[column].head(200)])
            sheet.column_dimensions[get_column_letter(col_index)].width = min(max(max_len * 1.7, 11), 42)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .stApp { background: linear-gradient(180deg, #F0F7FF 0, #FFFFFF 280px); }
        .block-container { max-width: 1500px; padding-top: 1.5rem; padding-bottom: 3rem; }
        .hero { padding: 1.7rem 2rem; border-radius: 20px; color: white;
                background: linear-gradient(120deg, #164E80, #2563EB); box-shadow: 0 12px 35px rgba(37,99,235,.18); }
        .hero h1 { margin: 0; font-size: 2rem; }
        .hero p { margin: .45rem 0 0; opacity: .9; }
        .section-label { color: #1E3A5F; font-size: 1.25rem; font-weight: 700; margin: 1.5rem 0 .5rem; }
        [data-testid="stMetric"] { background: white; border: 1px solid #DBEAFE; border-radius: 14px; padding: 1rem; box-shadow: 0 4px 18px rgba(15,23,42,.04); }
        .stButton > button { background: #2563EB; color: white; border: none; border-radius: 10px; font-weight: 700; min-height: 44px; }
        .stButton > button:hover { background: #1D4ED8; color: white; }
        .legend span { display:inline-block; padding:.25rem .55rem; margin:.15rem; border-radius:6px; font-size:.82rem; font-weight:600; }
        .shift-card { background:#FFFFFF; border:1px solid #DBEAFE; border-top:5px solid #2563EB;
                      border-radius:14px; padding:1rem 1.1rem; min-height:285px; margin-bottom:1rem;
                      box-shadow:0 5px 20px rgba(15,23,42,.06); }
        .shift-card.shortage { border-color:#FCA5A5; border-top-color:#DC2626; background:#FFF7F7; }
        .shift-card h3 { margin:0; color:#1E3A5F; font-size:1.15rem; }
        .shift-card .time { color:#64748B; font-size:.78rem; margin:.2rem 0 .65rem; }
        .shift-card .role { color:#1E40AF; font-weight:700; margin:.65rem 0 .25rem; }
        .shift-card ul { margin:.15rem 0 .35rem; padding-left:1.35rem; line-height:1.55; }
        .shift-card .badge { float:right; border-radius:999px; padding:.2rem .6rem; font-size:.75rem; font-weight:700;
                             background:#DCFCE7; color:#166534; }
        .shift-card .badge.shortage { background:#FEE2E2; color:#B91C1C; }
        .shift-card .badge.info { background:#E2E8F0; color:#475569; }
        .shift-card .staff-count { color:#64748B; font-size:.74rem; font-weight:500; }
        div[data-testid="stDataFrame"] { border: 1px solid #DBEAFE; border-radius: 12px; overflow: hidden; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🏥", layout="wide")
    inject_css()
    st.markdown(
        f'<div class="hero"><h1>🏥 {APP_TITLE}</h1><p>入所者100名規模の施設を想定した、説明可能な勤務表作成デモ</p></div>',
        unsafe_allow_html=True,
    )

    today = date.today()
    col_month, col_button, col_note = st.columns([1.3, 1, 3.5])
    with col_month:
        selected_month = st.date_input("対象月", value=date(today.year, today.month, 1), format="YYYY/MM/DD")
    with col_button:
        generate_clicked = st.button("✨ AIシフトを生成", use_container_width=True, type="primary")
    with col_note:
        st.caption("必要人数の充足、夜勤→明け→休み、連続勤務5日以内の順で自動配置します。")

    year, month = selected_month.year, selected_month.month
    key = f"{year}-{month:02d}"
    if generate_clicked:
        with st.spinner("勤務条件を確認し、1か月分を編成しています…"):
            st.session_state["schedule_result"] = generate_schedule(year, month)
            st.session_state["schedule_key"] = key

    if st.session_state.get("schedule_key") != key:
        st.info("対象月を選び、「AIシフトを生成」を押してください。ダミー職員38名も自動で用意されます。", icon="ℹ️")
        return

    result: GenerationResult = st.session_state["schedule_result"]
    shortage_count = int((result.daily["差分"] < 0).sum())
    total_shortage = int(-result.daily.loc[result.daily["差分"] < 0, "差分"].sum())
    nurse_nights = result.summary[result.summary["職種"] == "看護師"]["夜勤回数"].agg(["min", "max"])
    care_nights = result.summary[result.summary["職種"] == "介護士"]["夜勤回数"].agg(["min", "max"])

    st.markdown('<div class="section-label">編成サマリー</div>', unsafe_allow_html=True)
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("職員数", "38名", "看護師10・介護士28")
    m2.metric("対象期間", f"{year}年{month}月", f"{calendar.monthrange(year, month)[1]}日間")
    m3.metric("不足枠", f"{total_shortage}人日", f"{shortage_count}配置枠")
    m4.metric("看護師 夜勤", f"{int(nurse_nights['min'])}〜{int(nurse_nights['max'])}回", "公平性チェック")
    m5.metric("介護士 夜勤", f"{int(care_nights['min'])}〜{int(care_nights['max'])}回", "公平性チェック")
    if shortage_count == 0:
        st.success("全日・全勤務区分で必要人数を満たしています（不足0）。", icon="✅")

    st.markdown('<div class="section-label">シフト表示</div>', unsafe_allow_html=True)
    view_mode = st.radio(
        "表示形式",
        ["月間シフト表", "日別出勤者一覧"],
        horizontal=True,
        label_visibility="collapsed",
    )
    if view_mode == "月間シフト表":
        legend = "".join(f'<span style="background:{color}">{shift} {SHIFT_TIMES[shift]}</span>' for shift, color in SHIFT_COLORS.items())
        st.markdown(f'<div class="legend">{legend}</div>', unsafe_allow_html=True)
        schedule_view = display_schedule(result, year, month)
        st.dataframe(style_schedule(schedule_view), use_container_width=True, height=720, hide_index=True)
    else:
        day_count = calendar.monthrange(year, month)[1]
        selected_day = st.date_input(
            "表示する日付",
            value=date(year, month, min(date.today().day, day_count)) if (year, month) == (date.today().year, date.today().month) else date(year, month, 1),
            min_value=date(year, month, 1),
            max_value=date(year, month, day_count),
            format="YYYY/MM/DD",
            key=f"daily_roster_date_{year}_{month}",
        )
        st.markdown(
            f"### {selected_day.year}年{selected_day.month}月{selected_day.day}日（{WEEKDAYS[selected_day.weekday()]}）"
        )
        roster = _build_daily_roster(result.schedule, result.staff, [selected_day])
        card_columns = st.columns(3)
        for index, row in roster.iterrows():
            shift = row["勤務区分"]
            nurse_names = row["看護師名"].split("、") if row["看護師名"] != "—" else []
            care_names = row["介護士名"].split("、") if row["介護士名"] != "—" else []
            status_class = "shortage" if row["判定"] == "不足" else ("info" if row["判定"] == "確認用" else "")
            nurse_needed = REQUIRED["看護師"].get(shift)
            care_needed = REQUIRED["介護士"].get(shift)
            nurse_count = f"{len(nurse_names)}名" + (f" / 必要{nurse_needed}名" if nurse_needed is not None else "")
            care_count = f"{len(care_names)}名" + (f" / 必要{care_needed}名" if care_needed is not None else "")
            nurse_list = "".join(f"<li>{html.escape(name)}</li>" for name in nurse_names) or "<li>該当者なし</li>"
            care_list = "".join(f"<li>{html.escape(name)}</li>" for name in care_names) or "<li>該当者なし</li>"
            card_html = f"""
            <div class="shift-card {'shortage' if row['判定'] == '不足' else ''}">
              <span class="badge {status_class}">{row['判定']}</span>
              <h3>{html.escape(shift)}</h3>
              <div class="time">{html.escape(SHIFT_TIMES[shift])}</div>
              <div class="role">看護師 <span class="staff-count">{nurse_count}</span></div>
              <ul>{nurse_list}</ul>
              <div class="role">介護士 <span class="staff-count">{care_count}</span></div>
              <ul>{care_list}</ul>
            </div>
            """
            with card_columns[index % 3]:
                st.markdown(card_html, unsafe_allow_html=True)

    st.markdown('<div class="section-label">配置・集計</div>', unsafe_allow_html=True)
    tab_daily, tab_summary, tab_warning = st.tabs([
        "✅ 日別配置チェック", "👥 職員別集計", f"⚠️ 不足・警告（{len(result.warnings)}）"
    ])
    with tab_daily:
        daily_style = result.daily.style.map(
            lambda value: "background-color:#FEE2E2;color:#B91C1C;font-weight:700" if value == "不足" else "",
        )
        st.dataframe(daily_style, use_container_width=True, height=650, hide_index=True)
    with tab_summary:
        st.dataframe(result.summary, use_container_width=True, height=650, hide_index=True)
    with tab_warning:
        if (result.warnings["重要度"] == "正常").all():
            st.success("不足・警告はありません。")
        elif not result.warnings["重要度"].isin(["不足", "警告"]).any():
            st.warning(f"配置不足はありません。勤務条件に関する確認事項が {len(result.warnings)} 件あります。")
        else:
            st.error(f"不足または確認事項が {len(result.warnings)} 件あります。赤い項目をご確認ください。")
        warning_style = result.warnings.style.map(
            lambda value: "background-color:#FEE2E2;color:#B91C1C;font-weight:700" if value in ["不足", "警告"] else ""
        )
        st.dataframe(warning_style, use_container_width=True, height=600, hide_index=True)

    excel_bytes = make_excel(result, year, month)
    st.download_button(
        "📥 Excelファイルをダウンロード",
        data=excel_bytes,
        file_name=f"老健シフト表_{year}年{month:02d}月.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    with st.expander("このデモの編成ルール"):
        st.markdown(
            "- 最優先：全日・全勤務区分で必要人数を充足\n"
            "- 夜勤の翌日は必ず「明け」、その翌日は「休み」\n"
            "- 連続勤務は5日以内\n"
            "- 夜勤回数と土日勤務を職種内でできるだけ均等化\n"
            "- 月休日9日は目標値として確認事項に表示"
        )


if __name__ == "__main__":
    main()
